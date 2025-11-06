import os
import csv
import threading
import pandas as pd
import tkinter as tk
from datetime import datetime, timedelta
from tkinter import filedialog, messagebox, ttk
from utils import attendance_threshold, csv_file
from file_process import speak_notification, sort_column

def mark_attendance(name, mode="auto"):
    now = datetime.now()
    time_str = now.strftime('%Y-%m-%d %H:%M:%S')
    shift = determine_shift(now)
    shift_data = attendance_threshold.get(shift, None)
    if not shift_data:
        print(f"Shift {shift} tidak ditemukan dalam pengaturan!")
        return
    shift_hour = shift_data["hour"]
    shift_minute = shift_data["minute"]
    checkin_start_time = now.replace(hour=shift_hour - 1, minute=shift_minute, second=0, microsecond=0)
    checkin_end_time = now.replace(hour=shift_hour, minute=shift_minute, second=0, microsecond=0)
    late_checkin_end_time = checkin_end_time + timedelta(hours=4)
    if checkin_start_time <= now <= late_checkin_end_time:
        if now <= checkin_end_time + timedelta(minutes=5):
            status = "Tepat Waktu"
        else:
            status = "Terlambat"
        row_index, data = get_today_entry(name)
        if row_index is None:
            new_row = [name, time_str, status, '', shift]
            if not os.path.exists(csv_file):
                with open(csv_file, 'w', newline='') as f:
                    writer = csv.writer(f)
                    writer.writerow(['Name', 'TimeCheckIn', 'StatusCheckIn', 'TimeCheckOut', 'Shift'])
            with open(csv_file, 'a', newline='') as f:
                writer = csv.writer(f)
                writer.writerow(new_row)
            log_type = "Check-in"
        else:
            existing_row = data[row_index]
            check_in_time = datetime.strptime(existing_row[1], '%Y-%m-%d %H:%M:%S')
            if existing_row[3]:
                print(f"{name} sudah melakukan check-out hari ini.")
                return
            if now >= check_in_time + timedelta(minutes=1):
                data[row_index][3] = time_str
                with open(csv_file, 'w', newline='') as f:
                    writer = csv.writer(f)
                    writer.writerows(data)
                log_type = "Check-out"
            else:
                print(f"{name} sudah check-in, belum waktunya check-out.")
                return
    else:
        print(f"Waktu check-in untuk shift {shift} sudah lewat (lebih dari 4 jam).")
        return
    notification_message = f"{name} {log_type} ({shift}) - Status: {status}"
    print(notification_message)
    threading.Thread(target=speak_notification, args=(notification_message,)).start()
def show_today_attendance():
    if not os.path.exists(csv_file):
        messagebox.showerror("Error", "File absensi tidak ditemukan.")
        return
    today = datetime.now().strftime('%Y-%m-%d')
    with open(csv_file, 'r') as f:
        lines = f.readlines()
    if not lines or len(lines) == 1:
        messagebox.showinfo("Hari Ini", "Belum ada data absensi.")
        return
    headers = lines[0].strip().split(",")
    data = [line.strip().split(",") for line in lines[1:] if today in line]
    if not data:
        messagebox.showinfo("Hari Ini", "Belum ada yang hadir hari ini.")
        return
    top = tk.Toplevel()
    top.title(f"Data Absensi - {today}")
    top.geometry("700x450")
    top.configure(bg="#f0f0f0")
    label = ttk.Label(top, text=f"Daftar Kehadiran Hari Ini ({today})", font=('Segoe UI', 14, 'bold'))
    label.pack(pady=10)
    frame = ttk.Frame(top)
    frame.pack(fill="both", expand=True, padx=15, pady=10)
    tree = ttk.Treeview(frame, columns=headers, show="headings")
    tree.pack(side="left", fill="both", expand=True)
    scrollbar = ttk.Scrollbar(frame, orient="vertical", command=tree.yview)
    tree.configure(yscrollcommand=scrollbar.set)
    scrollbar.pack(side="right", fill="y")
    for header in headers:
        tree.heading(header, text=header, command=lambda _h=header: sort_column(tree, _h, False))
        tree.column(header, anchor='center', width=120, stretch=True)
    for i, row in enumerate(data):
        tag = 'evenrow' if i % 2 == 0 else 'oddrow'
        tree.insert('', 'end', values=row, tags=(tag,))
    style = ttk.Style()
    style.configure("Treeview.Heading", font=('Segoe UI', 10, 'bold'))
    style.configure("Treeview", font=('Segoe UI', 10), rowheight=25)
    style.map("Treeview", background=[("selected", "#d1e7dd")])
    tree.tag_configure('evenrow', background='#ffffff')
    tree.tag_configure('oddrow', background='#f5f5f5')
    def export_to_excel():
        file_path = filedialog.asksaveasfilename(
            defaultextension=".xlsx",
            filetypes=[("Excel files", "*.xlsx")],
            title="Simpan File Excel",
            initialfile=f"Absensi_{today}.xlsx"
        )
        if not file_path:
            return  
        df = pd.DataFrame(data, columns=headers)
        df.to_excel(file_path, index=False)
        from openpyxl import load_workbook
        from openpyxl.styles import Font, Alignment, PatternFill, Border, Side
        from openpyxl.utils import get_column_letter
        wb = load_workbook(file_path)
        ws = wb.active
        header_font = Font(bold=True, color="FFFFFF")
        header_fill = PatternFill("solid", fgColor="4F81BD") 
        align_center = Alignment(horizontal="center", vertical="center")
        thin_border = Border(
            left=Side(style='thin'),
            right=Side(style='thin'),
            top=Side(style='thin'),
            bottom=Side(style='thin')
        )
        for col_num, column_title in enumerate(headers, 1):
            cell = ws.cell(row=1, column=col_num)
            cell.font = header_font
            cell.fill = header_fill
            cell.alignment = align_center
            cell.border = thin_border
        for row in ws.iter_rows(min_row=2, max_row=ws.max_row, max_col=ws.max_column):
            for cell in row:
                cell.alignment = align_center
                cell.border = thin_border
        for col in ws.columns:
            max_length = max(len(str(cell.value)) if cell.value is not None else 0 for cell in col)
            adjusted_width = (max_length + 2)
            col_letter = get_column_letter(col[0].column)
            ws.column_dimensions[col_letter].width = adjusted_width
        ws.freeze_panes = ws['A2']
        wb.save(file_path)
        messagebox.showinfo("Berhasil", f"Data berhasil diekspor ke:\n{file_path}")
    export_btn = ttk.Button(top, text="Export ke Excel", command=export_to_excel)
    export_btn.pack(pady=5)
def determine_shift(current_time):
    if (current_time.hour >= attendance_threshold["pagi"]["hour"] - 1 and current_time.hour < attendance_threshold["siang"]["hour"] - 1):
        return "pagi"
    elif (current_time.hour >= attendance_threshold["siang"]["hour"] - 1 and current_time.hour < attendance_threshold["malam"]["hour"] - 1):
        return "siang"
    else:
        return "malam"
def get_today_entry(name):
    if not os.path.exists(csv_file):
        return None, []
    today = datetime.now().strftime('%Y-%m-%d')
    with open(csv_file, 'r', newline='') as f:
        reader = list(csv.reader(f))
    if not reader:
        return None, []
    header = reader[0]
    rows = reader[1:]

    for i, row in enumerate(rows):
        if row[0] == name and row[1].startswith(today):
            return i + 1, reader
    return None, reader