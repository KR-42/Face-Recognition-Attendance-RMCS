import cv2
import os
import pyttsx3
import logging
import json
from keras_facenet import FaceNet
from mtcnn import MTCNN
import threading
import numpy as np
from PIL import Image
import tkinter as tk
from tkinter import filedialog, messagebox, ttk
import ttkbootstrap as tb
from ttkbootstrap.constants import *
import csv
import pandas as pd
from datetime import datetime, timedelta
import openpyxl
from openpyxl.styles import PatternFill, Font, Alignment, Border, Side
from openpyxl.chart import BarChart, Reference
import hashlib
from concurrent.futures import ThreadPoolExecutor
import matplotlib.pyplot as plt
from matplotlib.backends.backend_tkagg import FigureCanvasTkAgg


embedder = FaceNet()
detector = MTCNN()
engine = pyttsx3.init()
notification_lock = threading.Lock()
engine.setProperty('rate', 150)
attendance_threshold = {"hour": 4}
password_file = "admin_password.json"
if os.path.exists("attendance_config.json"):
    with open("attendance_config.json", "r") as f:
        try:
            attendance_threshold = json.load(f)
        except json.JSONDecodeError:
            pass 
if os.path.exists(password_file):
    with open(password_file, "r") as f:
        try:
            admin_password = json.load(f).get("password", "admin123")
        except json.JSONDecodeError:
            admin_password = "admin123"
else:
    admin_password = "admin123"

dataset_dir = 'dataset_upg'
#dataset_dir = 'dataset1(FaceNet)'
dataset_tunggal = 'dataset_tunggal'
model_path = 'trainer.yml'
csv_file = 'attendance6.csv'
labels_file = 'labels1(FaceNet).txt'
log_file = 'app.log'
admin_file = 'admin_pass.txt'

os.makedirs(dataset_dir, exist_ok=True)
logging.basicConfig(filename=log_file, level=logging.INFO, format='%(asctime)s %(message)s')
face_cascade = cv2.CascadeClassifier(cv2.data.haarcascades + 'haarcascade_frontalface_default.xml')

def hash_password(password):
    return hashlib.sha256(password.encode()).hexdigest()
def get_today_entry(name):
    if not os.path.exists(csv_file):
        return None, []
    today = datetime.now().strftime('%Y-%m-%d')
    with open(csv_file, 'r', newline='') as f:
        reader = list(csv.reader(f))
    if not reader:
        return None, []
    header = reader[0]
    rows = reader[1:]
    for i, row in enumerate(rows):
        if row[0] == name and row[1].startswith(today):
            return i + 1, reader 
    return None, reader
def is_admin(username, password):
    if not os.path.exists(admin_file):
        return False
    hashed = hash_password(password)
    with open(admin_file, 'r') as f:
        for line in f:
            u, p = line.strip().split(',')
            if u == username and p == hashed:
                return True
    return False
def register_admin(username, password):
    with open(admin_file, 'a') as f:
        f.write(f"{username},{hash_password(password)}\n")
def speak_notification(message):
    with notification_lock:  
        engine.stop()  
        engine.say(message)  
        engine.runAndWait()
def get_today_entry(name):
    if not os.path.exists(csv_file):
        return None, []
    today = datetime.now().strftime('%Y-%m-%d')
    with open(csv_file, 'r', newline='') as f:
        reader = list(csv.reader(f))
    if not reader:
        return None, []
    header = reader[0]
    rows = reader[1:]

    for i, row in enumerate(rows):
        if row[0] == name and row[1].startswith(today):
            return i + 1, reader
    return None, reader
def mark_attendance(name, mode="auto"):
    now = datetime.now()
    time_str = now.strftime('%Y-%m-%d %H:%M:%S')
    shift = determine_shift(now)
    shift_data = attendance_threshold.get(shift, None)
    if not shift_data:
        print(f"Shift {shift} tidak ditemukan dalam pengaturan!")
        return
    shift_hour = shift_data["hour"]
    shift_minute = shift_data["minute"]
    checkin_start_time = now.replace(hour=shift_hour - 1, minute=shift_minute, second=0, microsecond=0)
    checkin_end_time = now.replace(hour=shift_hour, minute=shift_minute, second=0, microsecond=0)
    late_checkin_end_time = checkin_end_time + timedelta(hours=4)
    now = checkin_end_time - timedelta(minutes=1)
    if checkin_start_time <= now <= late_checkin_end_time:
        if now <= checkin_end_time + timedelta(minutes=5):
            status = "Tepat Waktu"
        else:
            status = "Terlambat"
        row_index, data = get_today_entry(name)
        if row_index is None:
            new_row = [name, time_str, status, '', shift]
            if not os.path.exists(csv_file):
                with open(csv_file, 'w', newline='') as f:
                    writer = csv.writer(f)
                    writer.writerow(['Name', 'TimeCheckIn', 'StatusCheckIn', 'TimeCheckOut', 'Shift'])
            with open(csv_file, 'a', newline='') as f:
                writer = csv.writer(f)
                writer.writerow(new_row)
            log_type = "Check-in"
        else:
            existing_row = data[row_index]
            check_in_time = datetime.strptime(existing_row[1], '%Y-%m-%d %H:%M:%S')
            if existing_row[3]:
                print(f"{name} sudah melakukan check-out hari ini.")
                return
            if now >= check_in_time + timedelta(minutes=1):
                data[row_index][3] = time_str
                with open(csv_file, 'w', newline='') as f:
                    writer = csv.writer(f)
                    writer.writerows(data)
                log_type = "Check-out"
            else:
                print(f"{name} sudah check-in, belum waktunya check-out.")
                return
    else:
        print(f"Waktu check-in untuk shift {shift} sudah lewat (lebih dari 4 jam).")
        return
    notification_message = f"{name} {log_type} ({shift}) - Status: {status}"
    print(notification_message)
    threading.Thread(target=speak_notification, args=(notification_message,)).start()
def determine_shift(current_time):
    if (current_time.hour >= attendance_threshold["pagi"]["hour"] - 1 and current_time.hour < attendance_threshold["siang"]["hour"] - 1):
        return "pagi"
    elif (current_time.hour >= attendance_threshold["siang"]["hour"] - 1 and current_time.hour < attendance_threshold["malam"]["hour"] - 1):
        return "siang"
    else:
        return "malam"
def detect_faces_mtcnn(image):
    results = detector.detect_faces(image)
    detected_faces = []
    for result in results:
        if result['confidence'] > 0.9:  
            detected_faces.append(result)
    return detected_faces
def register_face(name):
    if os.path.exists(os.path.join(dataset_dir, name)):
        messagebox.showerror("Error", "Nama sudah terdaftar. Gunakan nama lain.")
        return
    cam = cv2.VideoCapture(0)
    count = 0
    person_dir = os.path.join(dataset_dir, name)
    os.makedirs(person_dir, exist_ok=True)
    while True:
        ret, frame = cam.read()
        if not ret:
            break
        gray = cv2.cvtColor(frame, cv2.COLOR_BGR2GRAY)
        faces = face_cascade.detectMultiScale(gray, 1.3, 5)
        for (x, y, w, h) in faces:
            count += 1
            face_img = gray[y:y+h, x:x+w]
            cv2.imwrite(f"{person_dir}/{name}_{count}.jpg", face_img)
            cv2.rectangle(frame, (x,y), (x+w, y+h), (255,0,0), 2)
            cv2.putText(frame, f"{count}/40", (x, y-10), cv2.FONT_HERSHEY_SIMPLEX, 0.9, (255,255,255), 2)
        cv2.imshow('Register - Press q to quit', frame)
        if count >= 40 or cv2.waitKey(1) & 0xFF == ord('q'):
            break
    cam.release()
    cv2.destroyAllWindows()
    messagebox.showinfo("Done", f"Captured {count} faces for {name}")
    def save_embeddings_for_user(name):
        person_dir = os.path.join(dataset_dir, name)
        embeddings = []
        for image_name in os.listdir(person_dir):
            img_path = os.path.join(person_dir, image_name)
            img = cv2.imread(img_path)
            emb = get_face_embedding(img)
            if emb is not None:
                embeddings.append(emb)
        np.save(os.path.join(person_dir, f'{name}_embedding.npy'), embeddings)
    save_embeddings_for_user(name)
def train_model_from_folder(dataset_tunggal, dataset_dir):
    os.makedirs(dataset_dir, exist_ok=True)
    tasks = []
    with ThreadPoolExecutor(max_workers=8) as executor:
        for person_name in os.listdir(dataset_tunggal):
            person_dir = os.path.join(dataset_dir, person_name)
            image_path = os.path.join(dataset_tunggal, person_name, f"{person_name}.jpg")

            if not os.path.isfile(image_path):
                continue
            if os.path.exists(person_dir):  # skip jika sudah pernah dibuat
                continue

            img = cv2.imread(image_path)
            if img is None:
                print(f"Failed to read {image_path}")
                continue

            rgb = cv2.cvtColor(img, cv2.COLOR_BGR2RGB)
            # Submit task pendeteksian wajah
            future = executor.submit(detect_faces_mtcnn, rgb)
            tasks.append((future, person_name, rgb, person_dir))

        for future, person_name, rgb, person_dir in tasks:
            faces_results = future.result()

            if not faces_results:
                print(f"No face detected in {person_name}")
                continue

            os.makedirs(person_dir, exist_ok=True)

            # Proses satu wajah pertama
            for res in faces_results:
                x, y, w, h = res['box']
                x, y = max(0, x), max(0, y)
                face = rgb[y:y+h, x:x+w]
                face = cv2.resize(face, (160, 160))

                # Simpan 40 salinan wajah (bisa diganti dengan augmentasi)
                for i in range(1, 41):
                    file_name = f"{person_name}_{i}.jpg"
                    cv2.imwrite(os.path.join(person_dir, file_name), face)
                break  # hanya proses wajah pertama

            # Embedding
            embeddings = []
            for image_name in os.listdir(person_dir):
                img_path = os.path.join(person_dir, image_name)
                img = cv2.imread(img_path)
                emb = get_face_embedding(img)
                if emb is not None:
                    embeddings.append(emb)

            np.save(os.path.join(person_dir, f'{person_name}_embedding.npy'), embeddings)
def train_model():
    faces = []
    labels = []
    label_dict = {}
    current_id = 0
    valid_extensions = ('.jpg', '.jpeg', '.png', '.bmp') 
    for person_name in os.listdir(dataset_dir):
        person_dir = os.path.join(dataset_dir, person_name)
        if not os.path.isdir(person_dir):
            continue
        label_dict[current_id] = person_name
        for image_name in os.listdir(person_dir):
            if not image_name.lower().endswith(valid_extensions):
                continue  
            image_path = os.path.join(person_dir, image_name)
            try:
                img = Image.open(image_path).convert('L')
                img_np = np.array(img, 'uint8')
                faces.append(img_np)
                labels.append(current_id)
            except Exception as e:
                print(f"Error loading image {image_path}: {e}")
        current_id += 1
    with open(labels_file, 'w') as f:
        for id, name in label_dict.items():
            f.write(f"{id},{name}\n")
def train_model_from_lc():
    dataset_dir = 'dataset_upg'
    dataset_tunggal = 'dataset_tunggal'
    
    for filename in os.listdir(dataset_tunggal):
        image_path = os.path.join(dataset_tunggal, filename)
        if not os.path.isfile(image_path):
            continue

        person_name = filename.split('_')[0]
        person_dir = os.path.join(dataset_dir, person_name)
        if os.path.exists(person_dir):
            continue  # skip kalau folder sudah ada

        img = cv2.imread(image_path)
        if img is None:
            print(f"Failed to read {image_path}")
            continue

        rgb = cv2.cvtColor(img, cv2.COLOR_BGR2RGB)
        faces_results = detect_faces_mtcnn(rgb)

        if not faces_results:
            print(f"No face detected in {person_name}")
            continue

        # Ambil hanya wajah pertama
        x, y, w, h = faces_results[0]['box']
        face = img[y:y+h, x:x+w]

        os.makedirs(person_dir, exist_ok=True)

        # Simpan 40 kali dengan augmentasi ringan (opsional: rotasi, noise, dsb)
        for i in range(1, 41):
            file_name = f"{person_name}_{i}.jpg"
            save_path = os.path.join(person_dir, file_name)
            cv2.imwrite(save_path, face)

        # Buat embedding dari semua gambar yang disimpan
        embeddings = []
        for image_name in os.listdir(person_dir):
            img_path = os.path.join(person_dir, image_name)
            img = cv2.imread(img_path)
            emb = get_face_embedding(img)
            if emb is not None:
                embeddings.append(emb)

        np.save(os.path.join(person_dir, f'{person_name}_embedding.npy'), embeddings)
def load_labels():
    label_dict = {}
    if os.path.exists(labels_file):
        with open(labels_file, 'r') as f:
            for line in f:
                id, name = line.strip().split(',')
                label_dict[int(id)] = name
    return label_dict
def get_face_embedding(image):
    results = detect_faces_mtcnn(image)
    if results:
        x, y, w, h = results[0]['box']
        face = image[y:y+h, x:x+w]
        face = cv2.resize(face, (160, 160))
        face_rgb = cv2.cvtColor(face, cv2.COLOR_BGR2RGB)
        emb = embedder.embeddings([face_rgb])[0]
        return emb
    return None
def get_batch_face_embeddings(faces):
    embeddings = []
    faces_rgb = [cv2.cvtColor(face, cv2.COLOR_BGR2RGB) for face in faces]
    embeddings = embedder.embeddings(faces_rgb)  
    return embeddings
def recognize_faces_facenet_with_threads():
    data_vid = 'dataset4/vid_dua.mp4'
    if not os.path.exists(dataset_dir):
        messagebox.showerror("Error", "Belum ada data.")
        return
    cam = cv2.VideoCapture(0)
    known_embeddings = {}
    for person_name in os.listdir(dataset_dir):
        emb_path = os.path.join(dataset_dir, person_name, f'{person_name}_embedding.npy')
        if os.path.exists(emb_path):
            known_embeddings[person_name] = np.load(emb_path)
    recognized = []
    with ThreadPoolExecutor(max_workers=8) as executor: 
        while True:
            ret, frame = cam.read()
            if not ret:
                break
            rgb = cv2.cvtColor(frame, cv2.COLOR_BGR2RGB)
            futures = []
            future = executor.submit(detect_faces_mtcnn, rgb)
            faces_results = future.result()
            faces = []
            for res in faces_results:
                x, y, w, h = res['box']
                face = frame[y:y+h, x:x+w]
                faces.append(face)
            if faces:
                future_embeddings = executor.submit(get_batch_face_embeddings, faces)
                embeddings = future_embeddings.result()
                for emb, res in zip(embeddings, faces_results):
                    emb = emb / np.linalg.norm(emb)
                    x, y, w, h = res['box']
                    min_dist = float('inf')
                    identity = "Unknown"
                    for name, known_embs in known_embeddings.items():
                        for e in known_embs:
                            dist = np.linalg.norm(emb - e)
                            if dist < min_dist:
                                min_dist = dist
                                identity = name
                    if min_dist < 0.7:  
                        if identity not in recognized:
                            mark_attendance(identity, mode="auto")
                            recognized.append(identity)
                        color = (0, 255, 0)
                        label = f"{identity} ({min_dist:.2f})"
                    
                   
                    else:
                        color = (0, 0, 255)
                        label = "Unknown"
                    cv2.rectangle(frame, (x, y), (x+w, y+h), color, 2)
                    cv2.putText(frame, label, (x, y - 10), cv2.FONT_HERSHEY_SIMPLEX, 0.8, color, 2)
            cv2.imshow("Face Recognition - Press 'q' to Quit", frame)
            if cv2.waitKey(1) & 0xFF == ord('q'):
                break
    cam.release()
    cv2.destroyAllWindows()
def delete_user(name):
    person_dir = os.path.join(dataset_dir, name)
    if os.path.exists(person_dir):
        for file in os.listdir(person_dir):
            os.remove(os.path.join(person_dir, file))
        os.rmdir(person_dir)
        messagebox.showinfo("Delete", f"Data untuk {name} telah dihapus.")
    else:
        messagebox.showerror("Error", f"Data untuk {name} tidak ditemukan.")
def export_to_excel(root):
    def choose_save_location():
        file_path = filedialog.asksaveasfilename(defaultextension=".xlsx", filetypes=[("Excel files", "*.xlsx")])
        if file_path:
            try:
                df = pd.read_csv(csv_file)
                if df.empty:
                    messagebox.showinfo("Info", "Data absensi masih kosong.")
                    return
                start_date = start_date_entry.get().strip()
                end_date = end_date_entry.get().strip()
                try:
                    start_date = datetime.strptime(start_date, "%Y-%m-%d")
                    end_date = datetime.strptime(end_date, "%Y-%m-%d")
                except ValueError:
                    messagebox.showerror("Error", "Format tanggal tidak valid. Gunakan format YYYY-MM-DD.")
                    return
                df['TimeCheckIn'] = pd.to_datetime(df['TimeCheckIn'])
                filtered_df = df[(df['TimeCheckIn'] >= start_date) & (df['TimeCheckIn'] <= end_date)]
                if filtered_df.empty:
                    messagebox.showinfo("Info", "Tidak ada data untuk tanggal yang dipilih.")
                    return
                filtered_df['Weekday'] = filtered_df['TimeCheckIn'].dt.day_name()
                filtered_df.to_excel(file_path, index=False)
                wb = openpyxl.load_workbook(file_path)
                ws = wb.active
                ws.title = "Absensi"
                header_fill = PatternFill(start_color="4F81BD", end_color="4F81BD", fill_type="solid")
                header_font = Font(bold=True, size=12, color="FFFFFF")
                for col in range(1, len(filtered_df.columns) + 1):
                    cell = ws.cell(row=1, column=col)
                    cell.font = header_font
                    cell.alignment = Alignment(horizontal="center", vertical="center")
                    cell.fill = header_fill
                    cell.border = Border(
                        top=Side(style="thin"), bottom=Side(style="thin"),
                        left=Side(style="thin"), right=Side(style="thin")
                    )
                for row in range(2, ws.max_row + 1):
                    fill_color = PatternFill(start_color="FFFFFF", end_color="FFFFFF", fill_type="solid") if row % 2 == 0 else PatternFill(start_color="F2F2F2", end_color="F2F2F2", fill_type="solid")
                    for col in range(1, len(filtered_df.columns) + 1):
                        cell = ws.cell(row=row, column=col)
                        cell.fill = fill_color
                        cell.alignment = Alignment(horizontal="center", vertical="center")
                monday_fill = PatternFill(start_color="FFFF99", end_color="FFFF99", fill_type="solid")
                for row in range(2, ws.max_row + 1):  
                    weekday_cell = ws.cell(row=row, column=filtered_df.columns.get_loc('Weekday') + 1) 
                    if weekday_cell.value == "Monday":
                        for col in range(1, len(filtered_df.columns) + 1):
                            cell = ws.cell(row=row, column=col)
                            cell.fill = monday_fill
                for col in range(1, len(filtered_df.columns) + 1):
                    max_length = 0
                    column = chr(64 + col)
                    for row in ws[column]:
                        try:
                            if len(str(row.value)) > max_length:
                                max_length = len(row.value)
                        except:
                            pass
                    adjusted_width = (max_length + 2)
                    ws.column_dimensions[column].width = adjusted_width
                ws.freeze_panes = 'A2'
                tepat_waktu_count = filtered_df[filtered_df['StatusCheckIn'] == 'Tepat Waktu'].groupby('Name').size()
                terlambat_count = filtered_df[filtered_df['StatusCheckIn'] == 'Terlambat'].groupby('Name').size()
                chart_ws = wb.create_sheet(title="Perbandingan Grafik")
                chart_ws['A1'] = "Nama"
                chart_ws['B1'] = "Tepat Waktu"
                chart_ws['C1'] = "Terlambat"
                row = 2
                for name in set(tepat_waktu_count.index).union(set(terlambat_count.index)):
                    chart_ws.cell(row=row, column=1, value=name)
                    chart_ws.cell(row=row, column=2, value=tepat_waktu_count.get(name, 0))
                    chart_ws.cell(row=row, column=3, value=terlambat_count.get(name, 0))
                    row += 1
                chart = BarChart()
                data = Reference(chart_ws, min_col=2, min_row=1, max_col=3, max_row=row - 1)
                categories = Reference(chart_ws, min_col=1, min_row=2, max_row=row - 1)
                chart.add_data(data, titles_from_data=True)
                chart.set_categories(categories)
                chart.title = "Perbandingan Tepat Waktu dan Terlambat"
                chart.x_axis.title = "Nama"
                chart.y_axis.title = "Jumlah"
                chart.style = 10
                chart.width = 25
                chart.height = 15
                chart.series[0].graphicalProperties.line.solidFill = "4F81BD"
                chart.series[1].graphicalProperties.line.solidFill = "FF5733"
                chart_ws.add_chart(chart, "E5")
                wb.save(file_path)
                messagebox.showinfo("Sukses", f"Data berhasil diekspor ke {file_path}")
            except Exception as e:
                messagebox.showerror("Error", f"Terjadi kesalahan: {e}")
    date_window = tb.Toplevel(root)
    date_window.title("Pilih Rentang Tanggal")
    date_window.geometry("400x200")
    tb.Label(date_window, text="Masukkan tanggal mulai (YYYY-MM-DD):", bootstyle="info").pack(pady=5)
    start_date_entry = tb.Entry(date_window, width=30)
    start_date_entry.pack(pady=5)
    tb.Label(date_window, text="Masukkan tanggal akhir (YYYY-MM-DD):", bootstyle="info").pack(pady=5)
    end_date_entry = tb.Entry(date_window, width=30)
    end_date_entry.pack(pady=5)
    tb.Button(date_window, text="Pilih Lokasi dan Ekspor", command=choose_save_location, bootstyle="success").pack(pady=10)
def show_today_attendance():
    if not os.path.exists(csv_file):
        messagebox.showerror("Error", "File absensi tidak ditemukan.")
        return
    today = datetime.now().strftime('%Y-%m-%d')
    with open(csv_file, 'r') as f:
        lines = f.readlines()
    if not lines or len(lines) == 1:
        messagebox.showinfo("Hari Ini", "Belum ada data absensi.")
        return
    headers = lines[0].strip().split(",")
    data = [line.strip().split(",") for line in lines[1:] if today in line]
    if not data:
        messagebox.showinfo("Hari Ini", "Belum ada yang hadir hari ini.")
        return
    top = tk.Toplevel()
    top.title(f"Data Absensi - {today}")
    top.geometry("700x450")
    top.configure(bg="#f0f0f0")
    label = ttk.Label(top, text=f"Daftar Kehadiran Hari Ini ({today})", font=('Segoe UI', 14, 'bold'))
    label.pack(pady=10)
    frame = ttk.Frame(top)
    frame.pack(fill="both", expand=True, padx=15, pady=10)
    tree = ttk.Treeview(frame, columns=headers, show="headings")
    tree.pack(side="left", fill="both", expand=True)
    scrollbar = ttk.Scrollbar(frame, orient="vertical", command=tree.yview)
    tree.configure(yscrollcommand=scrollbar.set)
    scrollbar.pack(side="right", fill="y")
    for header in headers:
        tree.heading(header, text=header, command=lambda _h=header: sort_column(tree, _h, False))
        tree.column(header, anchor='center', width=120, stretch=True)
    for i, row in enumerate(data):
        tag = 'evenrow' if i % 2 == 0 else 'oddrow'
        tree.insert('', 'end', values=row, tags=(tag,))
    style = ttk.Style()
    style.configure("Treeview.Heading", font=('Segoe UI', 10, 'bold'))
    style.configure("Treeview", font=('Segoe UI', 10), rowheight=25)
    style.map("Treeview", background=[("selected", "#d1e7dd")])
    tree.tag_configure('evenrow', background='#ffffff')
    tree.tag_configure('oddrow', background='#f5f5f5')
    def export_to_excel():
        file_path = filedialog.asksaveasfilename(
            defaultextension=".xlsx",
            filetypes=[("Excel files", "*.xlsx")],
            title="Simpan File Excel",
            initialfile=f"Absensi_{today}.xlsx"
        )
        if not file_path:
            return  
        df = pd.DataFrame(data, columns=headers)
        df.to_excel(file_path, index=False)
        from openpyxl import load_workbook
        from openpyxl.styles import Font, Alignment, PatternFill, Border, Side
        from openpyxl.utils import get_column_letter
        wb = load_workbook(file_path)
        ws = wb.active
        header_font = Font(bold=True, color="FFFFFF")
        header_fill = PatternFill("solid", fgColor="4F81BD") 
        align_center = Alignment(horizontal="center", vertical="center")
        thin_border = Border(
            left=Side(style='thin'),
            right=Side(style='thin'),
            top=Side(style='thin'),
            bottom=Side(style='thin')
        )
        for col_num, column_title in enumerate(headers, 1):
            cell = ws.cell(row=1, column=col_num)
            cell.font = header_font
            cell.fill = header_fill
            cell.alignment = align_center
            cell.border = thin_border
        for row in ws.iter_rows(min_row=2, max_row=ws.max_row, max_col=ws.max_column):
            for cell in row:
                cell.alignment = align_center
                cell.border = thin_border
        for col in ws.columns:
            max_length = max(len(str(cell.value)) if cell.value is not None else 0 for cell in col)
            adjusted_width = (max_length + 2)
            col_letter = get_column_letter(col[0].column)
            ws.column_dimensions[col_letter].width = adjusted_width
        ws.freeze_panes = ws['A2']
        wb.save(file_path)
        messagebox.showinfo("Berhasil", f"Data berhasil diekspor ke:\n{file_path}")
    export_btn = ttk.Button(top, text="Export ke Excel", command=export_to_excel)
    export_btn.pack(pady=5)
def sort_column(tree, col, reverse):
    l = [(tree.set(k, col), k) for k in tree.get_children('')]
    try:
        l.sort(key=lambda t: float(t[0]) if t[0].replace('.', '', 1).isdigit() else t[0], reverse=reverse)
    except:
        l.sort(reverse=reverse)

    for index, (val, k) in enumerate(l):
        tree.move(k, '', index)
    tree.heading(col, command=lambda: sort_column(tree, col, not reverse))
def run_gui():
    def on_register():
        name = name_entry_user.get().strip()
        if name:
            register_face(name)
        else:
            messagebox.showerror("Error", "Isi nama dulu!")
    def on_train():
        train_model_from_lc()
        #train_model()
        #train_model_from_folder(dataset_tunggal, dataset_dir)
        messagebox.showinfo("Train", "Model training selesai!")
    def on_recognize():
        recognize_faces_facenet_with_threads()
    def on_delete():
        name = name_entry_admin.get().strip()
        if name:
            delete_user(name)
        else:
            messagebox.showerror("Error", "Masukkan nama untuk dihapus!")
    def on_export():
        export_to_excel()
    def on_today():
        show_today_attendance()
    def set_attendance_time():
        global attendance_threshold
        def save_time():
            try:
                pagi_hour = int(pagi_hour_combobox.get())
                pagi_minute = int(pagi_minute_combobox.get())
                siang_hour = int(siang_hour_combobox.get())
                siang_minute = int(siang_minute_combobox.get())
                malam_hour = int(malam_hour_combobox.get())
                malam_minute = int(malam_minute_combobox.get())
                if 0 <= pagi_hour <= 23 and 0 <= pagi_minute <= 59 and \
                0 <= siang_hour <= 23 and 0 <= siang_minute <= 59 and \
                0 <= malam_hour <= 23 and 0 <= malam_minute <= 59:
                    attendance_threshold["pagi"] = {"hour": pagi_hour, "minute": pagi_minute}
                    attendance_threshold["siang"] = {"hour": siang_hour, "minute": siang_minute}
                    attendance_threshold["malam"] = {"hour": malam_hour, "minute": malam_minute}
                    max_shift_hour = max(malam_hour, siang_hour, pagi_hour)
                    max_shift_minute = max(malam_minute, siang_minute, pagi_minute)
                    attendance_threshold["shift_start_time"] = {
                        "hour": (max_shift_hour - 1) % 24,
                        "minute": max_shift_minute
                    }
                    with open("attendance_config.json", "w") as f:
                        json.dump(attendance_threshold, f)
                    messagebox.showinfo("Berhasil", f"Pengaturan shift telah diset.")
                    config_window.destroy()
                else:
                    messagebox.showerror("Error", "Jam, menit tidak valid.")
            except ValueError:
                messagebox.showerror("Error", "Pilih jam, menit yang valid.")
        config_window = tb.Toplevel(root)
        config_window.title("Pengaturan Shift dan Jam Masuk")
        config_window.geometry("400x350")
        frame_main = tb.Frame(config_window, bootstyle="light")
        frame_main.pack(padx=20, pady=20, fill="both", expand=True)
        tb.Label(frame_main, text="Pengaturan Shift", font=("Segoe UI", 16, "bold"), bootstyle="info").pack(pady=10)
        tb.Label(frame_main, text="Pilih jam masuk shift pagi:", font=("Segoe UI", 12), bootstyle="secondary").pack(pady=5)
        pagi_hours = [str(i).zfill(2) for i in range(24)]
        pagi_hour_combobox = ttk.Combobox(frame_main, values=pagi_hours, state="readonly", font=("Segoe UI", 14))
        pagi_hour_combobox.set(str(attendance_threshold.get("pagi", {}).get("hour", 8)).zfill(2))
        pagi_hour_combobox.pack(pady=5, ipadx=5, ipady=5)
        pagi_minutes = [str(i).zfill(2) for i in range(60)]
        pagi_minute_combobox = ttk.Combobox(frame_main, values=pagi_minutes, state="readonly", font=("Segoe UI", 14))
        pagi_minute_combobox.set(str(attendance_threshold.get("pagi", {}).get("minute", 0)).zfill(2))
        pagi_minute_combobox.pack(pady=5, ipadx=5, ipady=5)
        tb.Label(frame_main, text="Pilih jam masuk shift siang:", font=("Segoe UI", 12), bootstyle="secondary").pack(pady=5)
        siang_hour_combobox = ttk.Combobox(frame_main, values=pagi_hours, state="readonly", font=("Segoe UI", 14))
        siang_hour_combobox.set(str(attendance_threshold.get("siang", {}).get("hour", 14)).zfill(2))
        siang_hour_combobox.pack(pady=5, ipadx=5, ipady=5)
        siang_minute_combobox = ttk.Combobox(frame_main, values=pagi_minutes, state="readonly", font=("Segoe UI", 14))
        siang_minute_combobox.set(str(attendance_threshold.get("siang", {}).get("minute", 0)).zfill(2))
        siang_minute_combobox.pack(pady=5, ipadx=5, ipady=5)
        tb.Label(frame_main, text="Pilih jam masuk shift malam:", font=("Segoe UI", 12), bootstyle="secondary").pack(pady=5)
        malam_hour_combobox = ttk.Combobox(frame_main, values=pagi_hours, state="readonly", font=("Segoe UI", 14))
        malam_hour_combobox.set(str(attendance_threshold.get("malam", {}).get("hour", 22)).zfill(2))
        malam_hour_combobox.pack(pady=5, ipadx=5, ipady=5)
        malam_minute_combobox = ttk.Combobox(frame_main, values=pagi_minutes, state="readonly", font=("Segoe UI", 14))
        malam_minute_combobox.set(str(attendance_threshold.get("malam", {}).get("minute", 0)).zfill(2))
        malam_minute_combobox.pack(pady=5, ipadx=5, ipady=5)
        tb.Button(frame_main, text="Simpan", command=save_time, bootstyle="success", width=20).pack(pady=10)
        tb.Button(frame_main, text="Batal", command=config_window.destroy, bootstyle="danger", width=20).pack(pady=5)
    def show_salary_window():
        salary_window = tb.Toplevel()
        salary_window.title("Salary Calculation")
        salary_window.geometry("700x500")
        salary_window.resizable(False, False)
        input_frame = tb.LabelFrame(salary_window, text="Input Data Karyawan", bootstyle="primary")
        input_frame.pack(fill="x", padx=20, pady=10)
        tb.Label(input_frame, text="Nama Karyawan:", bootstyle="info").grid(row=0, column=0, sticky="w", padx=10, pady=5)
        name_entry = tb.Entry(input_frame, width=30)
        name_entry.grid(row=0, column=1, padx=10, pady=5)
        tb.Label(input_frame, text="Gaji per Jam (Rp):", bootstyle="info").grid(row=1, column=0, sticky="w", padx=10, pady=5)
        wage_entry = tb.Entry(input_frame, width=30)
        wage_entry.grid(row=1, column=1, padx=10, pady=5)
        tb.Label(input_frame, text="Tanggal Mulai (YYYY-MM-DD):", bootstyle="info").grid(row=2, column=0, sticky="w", padx=10, pady=5)
        start_entry = tb.Entry(input_frame, width=30)
        start_entry.grid(row=2, column=1, padx=10, pady=5)
        tb.Label(input_frame, text="Tanggal Akhir (YYYY-MM-DD):", bootstyle="info").grid(row=3, column=0, sticky="w", padx=10, pady=5)
        end_entry = tb.Entry(input_frame, width=30)
        end_entry.grid(row=3, column=1, padx=10, pady=5)
        result_frame = tb.Frame(salary_window)
        result_frame.pack(pady=10, fill="both", expand=True)
        def calculate_salary():
            name = name_entry.get().strip()
            wage = wage_entry.get().strip()
            start = start_entry.get().strip()
            end = end_entry.get().strip()
            if not name or not wage or not start or not end:
                messagebox.showwarning("Warning", "Isi semua data terlebih dahulu.")
                return
            try:
                hourly_rate = float(wage)
                start_date = datetime.strptime(start, "%Y-%m-%d")
                end_date = datetime.strptime(end, "%Y-%m-%d")
            except ValueError:
                messagebox.showerror("Error", "Masukkan data yang valid.")
                return
            try:
                df = pd.read_csv(csv_file)
                df.columns = df.columns.str.strip()
                df = df[df['Name'].str.strip().str.lower() == name.lower()]
                if df.empty:
                    messagebox.showerror("Error", f"Nama {name} tidak ditemukan.")
                    return
                df['TimeCheckIn'] = pd.to_datetime(df['TimeCheckIn'])
                df['TimeCheckOut'] = pd.to_datetime(df['TimeCheckOut'], errors='coerce')
                df = df[(df['TimeCheckIn'] >= start_date) & (df['TimeCheckIn'] <= end_date)]
                df['Duration'] = (df['TimeCheckOut'] - df['TimeCheckIn']).dt.total_seconds() / 3600
                df['Duration'] = df['Duration'].fillna(0).clip(lower=0)
                df['Gaji'] = df['Duration'] * hourly_rate
                total_salary = df['Gaji'].sum()
                for widget in result_frame.winfo_children():
                    widget.destroy()
                tb.Label(result_frame, text=f"💼 Total Gaji: Rp {total_salary:,.2f}",
                        bootstyle="success", font=("Helvetica", 13, "bold")).pack(pady=10)
                columns = ['Name', 'TimeCheckIn', 'TimeCheckOut', 'StatusCheckIn', 'Duration', 'Gaji']
                tree_frame = tb.Frame(result_frame)
                tree_frame.pack(fill="both", expand=True)
                tree_scroll = ttk.Scrollbar(tree_frame)
                tree_scroll.pack(side="right", fill="y")
                tree = ttk.Treeview(tree_frame, columns=columns, show="headings", yscrollcommand=tree_scroll.set)
                tree_scroll.config(command=tree.yview)
                for col in columns:
                    tree.heading(col, text=col)
                    tree.column(col, width=110, anchor="center")
                for _, row in df.iterrows():
                    tree.insert("", "end", values=[
                        row['Name'],
                        row['TimeCheckIn'].strftime("%Y-%m-%d %H:%M:%S") if pd.notnull(row['TimeCheckIn']) else '',
                        row['TimeCheckOut'].strftime("%Y-%m-%d %H:%M:%S") if pd.notnull(row['TimeCheckOut']) else '',
                        row.get('StatusCheckIn', ''),
                        f"{row['Duration']:.2f}",
                        f"Rp {row['Gaji']:.2f}"
                    ])
                tree.pack(fill="both", expand=True)
            except Exception as e:
                messagebox.showerror("Error", f"Terjadi kesalahan saat menghitung gaji:\n{e}")
        tb.Button(salary_window, text="🔍 Calculate", command=calculate_salary,
                bootstyle="primary outline").pack(pady=10)
    def show_all_attendance():
        if not os.path.exists(csv_file):
            messagebox.showerror("Error", "File absensi tidak ditemukan.")
            return
        df = pd.read_csv(csv_file)
        if df.empty:
            messagebox.showinfo("Absensi", "Data absensi masih kosong.")
            return
        top = tk.Toplevel(root)
        top.title("Semua Data Absensi")
        top.geometry("1000x550")
        top.configure(bg="#f0f0f0")
        title = ttk.Label(top, text="Data Absensi Keseluruhan", font=('Segoe UI', 14, 'bold'))
        search_frame = ttk.Frame(top)
        search_frame.pack(pady=5)
        search_label = ttk.Label(search_frame, text="Cari Nama atau Tanggal:")
        search_label.pack(side="left", padx=(0, 5))
        search_var = tk.StringVar()
        search_entry = ttk.Entry(search_frame, textvariable=search_var, width=30)
        search_entry.pack(side="left")
        def filter_data():
            keyword = search_var.get().lower().strip()
            filtered_df = df[df.apply(lambda row: row.astype(str).str.lower().str.contains(keyword).any(), axis=1)]
            for item in tree.get_children():
                tree.delete(item)
            for i, row in filtered_df.iterrows():
                tag = 'evenrow' if i % 2 == 0 else 'oddrow'
                tree.insert('', 'end', values=list(row), tags=(tag,))
        search_btn = ttk.Button(search_frame, text="Cari", command=filter_data)
        search_btn.pack(side="left", padx=5)
        reset_btn = ttk.Button(search_frame, text="Reset", command=lambda: [search_var.set(""), populate_tree(df)])
        reset_btn.pack(side="left")
        frame = ttk.Frame(top)
        frame.pack(fill="both", expand=True, padx=15, pady=10)
        tree = ttk.Treeview(frame, columns=list(df.columns), show="headings")
        tree.pack(side="left", fill="both", expand=True)
        scrollbar = ttk.Scrollbar(frame, orient="vertical", command=tree.yview)
        tree.configure(yscrollcommand=scrollbar.set)
        scrollbar.pack(side="right", fill="y")
        for col in df.columns:
            tree.heading(col, text=col, anchor='center')
            tree.column(col, anchor='center', width=120, stretch=True)
        def populate_tree(dataframe):
            for item in tree.get_children():
                tree.delete(item)
            for i, row in dataframe.iterrows():
                tag = 'evenrow' if i % 2 == 0 else 'oddrow'
                tree.insert('', 'end', values=list(row), tags=(tag,))
        populate_tree(df)
        style = ttk.Style()
        style.configure("Treeview.Heading", font=('Segoe UI', 10, 'bold'))
        style.configure("Treeview", font=('Segoe UI', 10), rowheight=25)
        style.map("Treeview", background=[("selected", "#d1e7dd")])
        tree.tag_configure('evenrow', background='#ffffff')
        tree.tag_configure('oddrow', background='#f5f5f5')
    def switch_to_admin():
        def verify_password():
            entered_password = password_entry.get()
            if entered_password == admin_password:
                messagebox.showinfo("Akses Diberikan", "Berhasil masuk sebagai admin.")
                notebook.add(admin_tab, text="Admin")
                notebook.select(admin_tab)
                pw_window.destroy()
            else:
                messagebox.showerror("Akses Ditolak", "Password salah!")
        pw_window = tb.Toplevel(root)
        pw_window.title("Verifikasi Admin")
        pw_window.geometry("300x150")
        tb.Label(pw_window, text="Masukkan Password Admin:", bootstyle="info").pack(pady=10)
        password_entry = tb.Entry(pw_window, show="*", width=25)
        password_entry.pack(pady=5)
        tb.Button(pw_window, text="Masuk", command=verify_password, bootstyle="primary").pack(pady=10)
        password_entry.focus()
    def change_admin_password():
        def verify_old_password():
            if old_pass_entry.get() == admin_password:
                old_pass_window.destroy()
                enter_new_password()
            else:
                messagebox.showerror("Error", "Password lama salah!")
        def enter_new_password():
            def save_new_password():
                new_pass = new_pass_entry.get()
                if len(new_pass) < 4:
                    messagebox.showerror("Error", "Password terlalu pendek (min. 4 karakter)")
                    return
                with open(password_file, "w") as f:
                    json.dump({"password": new_pass}, f)
                messagebox.showinfo("Berhasil", "Password admin berhasil diganti!")
                new_pass_window.destroy()
                global admin_password
                admin_password = new_pass
            new_pass_window = tb.Toplevel(root)
            new_pass_window.title("Password Baru")
            new_pass_window.geometry("300x150")
            tb.Label(new_pass_window, text="Masukkan Password Baru:", bootstyle="info").pack(pady=10)
            new_pass_entry = tb.Entry(new_pass_window, show="*", width=25)
            new_pass_entry.pack(pady=5)
            tb.Button(new_pass_window, text="Simpan", command=save_new_password, bootstyle="success").pack(pady=10)
        old_pass_window = tb.Toplevel(root)
        old_pass_window.title("Verifikasi Password Lama")
        old_pass_window.geometry("300x150")
        tb.Label(old_pass_window, text="Masukkan Password Lama:", bootstyle="warning").pack(pady=10)
        old_pass_entry = tb.Entry(old_pass_window, show="*", width=25)
        old_pass_entry.pack(pady=5)
        tb.Button(old_pass_window, text="Lanjut", command=verify_old_password, bootstyle="primary").pack(pady=10)
    def filter_by_name():
        def search():
            name = name_entry_filter.get().strip()
            if os.path.exists(csv_file) and name:
                df = pd.read_csv(csv_file)
                filtered = df[df["Name"].str.lower() == name.lower()]
                if not filtered.empty:
                    populate_tree(filtered)
                else:
                    for item in tree.get_children():
                        tree.delete(item)
                    messagebox.showinfo("Info", f"Tidak ditemukan data untuk '{name}'")
            else:
                messagebox.showerror("Error", "Masukkan nama yang ingin dicari.")
        filter_window = tb.Toplevel(root)
        filter_window.title("Riwayat Kehadiran Pegawai")
        filter_window.geometry("900x500")
        filter_window.configure(bg="#f0f0f0")
        tb.Label(filter_window, text="Masukkan Nama:", bootstyle="info").pack(pady=5)
        name_entry_filter = tb.Entry(filter_window, width=30)
        name_entry_filter.pack(pady=5)
        tb.Button(filter_window, text="Cari", command=search, bootstyle="primary").pack(pady=5)
        frame = ttk.Frame(filter_window)
        frame.pack(fill="both", expand=True, padx=15, pady=10)
        tree = ttk.Treeview(frame, show="headings")
        tree.pack(side="left", fill="both", expand=True)
        scrollbar = ttk.Scrollbar(frame, orient="vertical", command=tree.yview)
        tree.configure(yscrollcommand=scrollbar.set)
        scrollbar.pack(side="right", fill="y")
        style = ttk.Style()
        style.configure("Treeview.Heading", font=('Segoe UI', 10, 'bold'))
        style.configure("Treeview", font=('Segoe UI', 10), rowheight=25)
        style.map("Treeview", background=[("selected", "#d1e7dd")])
        tree.tag_configure('evenrow', background='#ffffff')
        tree.tag_configure('oddrow', background='#f5f5f5')
        def populate_tree(dataframe):
            tree["columns"] = list(dataframe.columns)
            for col in dataframe.columns:
                tree.heading(col, text=col, anchor='center')
                tree.column(col, anchor='center', width=120, stretch=True)
            for item in tree.get_children():
                tree.delete(item)
            for i, row in dataframe.iterrows():
                tag = 'evenrow' if i % 2 == 0 else 'oddrow'
                tree.insert('', 'end', values=list(row), tags=(tag,))
    def reset_attendance():
        if os.path.exists(csv_file):
            if messagebox.askyesno("Konfirmasi", "Yakin ingin menghapus seluruh data absensi?"):
                os.remove(csv_file)
                messagebox.showinfo("Reset", "Data absensi telah dihapus.")
        else:
            messagebox.showerror("Error", "File absensi tidak ditemukan.")
    def show_graph_window():
        if not os.path.exists(csv_file):
            messagebox.showerror("Error", "Belum ada data absensi.")
            return
        df = pd.read_csv(csv_file)
        if df.empty:
            messagebox.showinfo("Info", "Data absensi masih kosong.")
            return
        if 'StatusCheckIn' not in df.columns:
            messagebox.showerror("Error", "Kolom 'StatusCheckIn' tidak ditemukan.")
            return
        def plot_graph(name=None):
            ax.clear()
            if name:
                person_data = df[df["Name"].str.lower() == name.lower()]
                if person_data.empty:
                    messagebox.showinfo("Info", f"Tidak ada data untuk {name}")
                    return
                count_tepat = (person_data["StatusCheckIn"] == "Tepat Waktu").sum()
                count_terlambat = (person_data["StatusCheckIn"] == "Terlambat").sum()
                title = f"Grafik Kehadiran - {name.title()}"
            else:
                count_tepat = (df["StatusCheckIn"] == "Tepat Waktu").sum()
                count_terlambat = (df["StatusCheckIn"] == "Terlambat").sum()
                title = "Grafik Kehadiran (Semua)"
            bars = ax.bar(
                ["Tepat Waktu", "Terlambat"],
                [count_tepat, count_terlambat],
                color=["#4CAF50", "#F44336"],
                edgecolor='black'
            )
            for bar in bars:
                height = bar.get_height()
                ax.annotate(f'{height}',
                            xy=(bar.get_x() + bar.get_width() / 2, height),
                            xytext=(0, 5),
                            textcoords="offset points",
                            ha='center', va='bottom', fontsize=10, fontweight='bold')

            ax.set_title(title, fontsize=14, fontweight='bold')
            ax.set_ylabel("Jumlah", fontsize=12)
            ax.set_xlabel("Status", fontsize=12)
            ax.grid(axis='y', linestyle='--', alpha=0.7)
            canvas.draw()
        def on_combo_select(event):
            selected = combo_name.get()
            if selected == "(Semua)":
                plot_graph()
            else:
                plot_graph(selected)
        def on_search():
            search_name = entry_search.get().strip()
            if not search_name:
                messagebox.showerror("Error", "Masukkan nama untuk dicari.")
                return
            plot_graph(search_name)
        graph_win = tb.Toplevel(root)
        graph_win.title("Grafik Kehadiran")
        graph_win.geometry("720x560")
        graph_win.configure(bg="#f8f9fa")
        frame_top = tb.Frame(graph_win)
        frame_top.pack(pady=10)
        tb.Label(frame_top, text="Pilih Nama:", bootstyle="secondary").grid(row=0, column=0, padx=5)
        names = sorted(df["Name"].unique().tolist())
        combo_name = ttk.Combobox(frame_top, values=["(Semua)"] + names, width=30)
        combo_name.set("(Semua)")
        combo_name.grid(row=0, column=1, padx=5)
        combo_name.bind("<<ComboboxSelected>>", on_combo_select)
        tb.Label(frame_top, text="Atau cari nama:", bootstyle="secondary").grid(row=1, column=0, padx=5, pady=5)
        entry_search = tb.Entry(frame_top, width=33)
        entry_search.grid(row=1, column=1, padx=5, pady=5)
        tb.Button(frame_top, text="Cari", command=on_search, bootstyle="info").grid(row=1, column=2, padx=5)
        fig, ax = plt.subplots(figsize=(6, 4))
        fig.tight_layout(pad=4.0)
        canvas = FigureCanvasTkAgg(fig, master=graph_win)
        canvas.get_tk_widget().pack(fill=tk.BOTH, expand=True, padx=10, pady=10)
        plot_graph()
    
    root = tb.Window(themename="flatly")
    root.title("Face Attendance (Extended)")
    root.geometry("480x740")
    notebook = tb.Notebook(root)
    notebook.pack(fill='both', expand=True, padx=10, pady=10)
    user_tab = tb.Frame(notebook)
    admin_tab = tb.Frame(notebook)
    notebook.add(user_tab, text="User")
    tb.Label(user_tab, text="Face Recognition Attendance", font=("Arial", 16, "bold"), bootstyle="primary").pack(pady=20)
    tb.Button(user_tab, text="Mulai Absensi", command=on_recognize, bootstyle="success", width=25).pack(pady=5)
    tb.Button(user_tab, text="Pindah ke Admin", command=switch_to_admin, bootstyle="secondary", width=25).pack(pady=15)
    tb.Label(admin_tab, text="Menu Admin", font=("Arial", 16, "bold"), bootstyle="dark").pack(pady=20)
    tb.Label(admin_tab, text="Nama untuk dihapus:", bootstyle="info").pack()
    name_entry_admin = tb.Entry(admin_tab, width=30)
    name_entry_admin.pack(pady=5)
    tb.Label(admin_tab, text="Registrasi Wajah:", bootstyle="info").pack(pady=5)
    name_entry_user = tb.Entry(admin_tab, width=30)
    name_entry_user.pack(pady=5)
    tb.Button(admin_tab, text="Register Wajah", command=on_register, bootstyle="success", width=25).pack(pady=5)
    tb.Button(admin_tab, text="Train Model", command=on_train, bootstyle="success", width=25).pack(pady=5)
    tb.Button(admin_tab, text="Hapus Data Wajah", command=on_delete, bootstyle="danger", width=25).pack(pady=5)
    tb.Button(admin_tab, text="Export ke Excel", command=lambda: export_to_excel(root), bootstyle="info", width=25).pack(pady=5)
    tb.Button(admin_tab, text="Lihat Kehadiran Hari Ini", command=on_today, bootstyle="secondary", width=25).pack(pady=5)
    tb.Button(admin_tab, text="Ganti Password Admin", command=change_admin_password, bootstyle="warning").pack(pady=10)
    tb.Button(admin_tab, text="Set Jam Masuk", command=set_attendance_time, bootstyle="warning", width=25).pack(pady=5)
    tb.Button(admin_tab, text="Lihat Semua Absensi", command=show_all_attendance, width=25).pack(pady=5)
    tb.Button(admin_tab, text="Filter Absensi Berdasarkan Nama", command=filter_by_name, bootstyle="primary", width=25).pack(pady=5)
    tb.Button(admin_tab, text="Reset Semua Absensi", command=reset_attendance, bootstyle="danger", width=25).pack(pady=5)
    tb.Button(admin_tab, text="Salary Calculation", command=show_salary_window, bootstyle="secondary", width=25).pack(pady=5)
    tb.Button(admin_tab, text="Lihat Grafik Absensi", command=show_graph_window, bootstyle="info", width=25).pack(pady=5)
    tb.Button(admin_tab, text="Keluar", command=root.quit, bootstyle="dark", width=25).pack(pady=15)
    root.mainloop()
if __name__ == "__main__":
    run_gui()